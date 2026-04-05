"""
导入解析服务模块

提供 task_params 与 flow_input 共用的 JSON、店铺标识、CSV/XLSX 解析能力。
"""
from __future__ import annotations

import csv
import io
import json
import re
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

from backend.models.database import 获取连接


科学计数法正则 = re.compile(r"^[+-]?\d+\.?\d*[eE][+\-]?\d+$")
整数值正则 = re.compile(r"^\d+$")
整数浮点正则 = re.compile(r"^\d+\.0+$")
参数键名映射 = {
    "父商品ID": "parent_product_id",
    "parent_product_id": "parent_product_id",
    "新标题": "new_title",
    "new_title": "new_title",
    "图片路径": "image_path",
    "image_path": "image_path",
    "批次ID": "batch_id",
    "批次号": "batch_id",
    "批次编号": "batch_id",
    "batch_id": "batch_id",
    "batchId": "batch_id",
    "折扣": "discount",
    "discount": "discount",
    "投产比": "roi",
    "roi": "roi",
}
保持字符串字段 = {
    "parent_product_id",
    "new_title",
    "image_path",
    "batch_id",
}


class 导入解析服务:
    """封装导入链路共用的解析能力。"""

    @staticmethod
    def 序列化JSON(数据: Optional[Dict[str, Any]]) -> str:
        return json.dumps(数据 or {}, ensure_ascii=False)

    @staticmethod
    def 解析JSON(数据: Any) -> Dict[str, Any]:
        if not 数据:
            return {}
        if isinstance(数据, dict):
            return 数据
        try:
            return json.loads(数据)
        except (TypeError, json.JSONDecodeError):
            return {}

    async def 店铺是否存在(self, 店铺ID: str) -> bool:
        async with 获取连接() as 连接:
            async with 连接.execute(
                "SELECT 1 FROM shops WHERE id = ?",
                (店铺ID,),
            ) as 游标:
                return await 游标.fetchone() is not None

    async def 根据店铺名称获取ID(self, 店铺名称: str) -> Optional[str]:
        async with 获取连接() as 连接:
            async with 连接.execute(
                "SELECT id FROM shops WHERE name = ? ORDER BY created_at DESC LIMIT 1",
                (店铺名称,),
            ) as 游标:
                结果 = await 游标.fetchone()
                if not 结果:
                    return None
                return str(结果["id"])

    async def 解析店铺标识(self, 店铺标识: str, 行号: int) -> str:
        if 店铺标识.isdigit():
            if not await self.店铺是否存在(店铺标识):
                raise ValueError(f"第 {行号} 行店铺ID不存在: {店铺标识}")
            return 店铺标识

        if await self.店铺是否存在(店铺标识):
            return 店铺标识

        店铺ID = await self.根据店铺名称获取ID(店铺标识)
        if 店铺ID:
            return 店铺ID

        raise ValueError(f"第 {行号} 行店铺名称未找到: {店铺标识}")

    @staticmethod
    def 解码CSV文本(文件内容: bytes) -> str:
        for 编码 in ("utf-8-sig", "utf-8", "gbk"):
            try:
                return 文件内容.decode(编码)
            except (UnicodeDecodeError, ValueError):
                continue
        raise ValueError("CSV 文件编码不支持，请另存为 UTF-8 格式")

    @staticmethod
    def 修复科学计数法(值: str) -> str:
        标准值 = str(值).strip()
        if not 标准值:
            return 标准值
        if 科学计数法正则.match(标准值):
            return str(int(float(标准值)))
        return 标准值

    def 预处理CSV行(self, 行数据: Dict[str, str]) -> Dict[str, str]:
        结果: Dict[str, str] = {}
        for 列名, 原值 in 行数据.items():
            字段值 = "" if 原值 is None else str(原值)
            if 列名 and any(标记 in 列名 for 标记 in ("ID", "id", "Id")):
                结果[列名] = self.修复科学计数法(字段值)
            else:
                结果[列名] = 字段值
        return 结果

    def 解析CSV内容(self, 文件内容: bytes) -> List[Dict[str, str]]:
        文本内容 = self.解码CSV文本(文件内容)
        读取器 = csv.DictReader(io.StringIO(文本内容))
        if not 读取器.fieldnames:
            raise ValueError("CSV 文件缺少表头")
        return [self.预处理CSV行(dict(行)) for 行 in 读取器]

    def 解析XLSX内容(self, 文件内容: bytes) -> List[Dict[str, str]]:
        """读取 xlsx 内容并转换为与 CSV 一致的行结构。"""
        工作簿 = load_workbook(io.BytesIO(文件内容), read_only=True)
        try:
            工作表 = 工作簿.worksheets[0]
            行迭代器 = 工作表.iter_rows()
            表头行 = next(行迭代器, None)
            if not 表头行:
                raise ValueError("XLSX 文件缺少表头")

            表头 = ["" if 单元格.value is None else str(单元格.value).strip() for 单元格 in 表头行]
            if not any(表头):
                raise ValueError("XLSX 文件缺少表头")

            结果列表: List[Dict[str, str]] = []
            for 数据行 in 行迭代器:
                行数据: Dict[str, str] = {}
                for 索引, 单元格 in enumerate(数据行):
                    列名 = 表头[索引] if 索引 < len(表头) else ""
                    if not 列名:
                        continue

                    if 单元格.value is None:
                        行数据[列名] = ""
                        continue

                    if 单元格.data_type == "n" and 单元格.value > 9999999999:
                        行数据[列名] = str(int(单元格.value))
                    else:
                        行数据[列名] = str(单元格.value)

                if any(str(值).strip() for 值 in 行数据.values()):
                    结果列表.append(self.预处理CSV行(行数据))

            return 结果列表
        finally:
            工作簿.close()

    @staticmethod
    def 解析发布次数(行数据: Dict[str, str], 行号: int) -> int:
        原始值 = str(行数据.get("发布次数", "")).strip()
        if not 原始值:
            return 1

        if 整数值正则.match(原始值):
            发布次数 = int(原始值)
        elif 整数浮点正则.match(原始值):
            发布次数 = int(float(原始值))
        else:
            raise ValueError(f"第 {行号} 行发布次数必须是正整数")

        if 发布次数 <= 0:
            raise ValueError(f"第 {行号} 行发布次数必须大于 0")
        return 发布次数

    @staticmethod
    def 规范参数键名(列名: str) -> str:
        标准列名 = str(列名).strip()
        return 参数键名映射.get(标准列名, 标准列名)

    @staticmethod
    def 转换参数值(参数键名: str, 原始值: str) -> Any:
        清理值 = str(原始值).strip()
        if not 清理值:
            return 清理值

        if 参数键名 in 保持字符串字段:
            return 清理值

        if 参数键名.endswith("_id"):
            return 清理值

        try:
            return float(清理值) if "." in 清理值 else int(清理值)
        except ValueError:
            return 清理值


导入解析服务实例 = 导入解析服务()

