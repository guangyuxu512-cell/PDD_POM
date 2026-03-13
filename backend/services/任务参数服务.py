"""
任务参数服务模块

提供 task_params 表的分页查询、CRUD、CSV 批量导入与执行状态更新能力。
"""
import csv
import io
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from backend.models.数据库 import 获取连接


允许状态集合 = {"pending", "running", "success", "failed", "skipped"}
CSV映射配置: Dict[str, Dict[str, str]] = {
    "发布相似商品": {
        "店铺ID": "shop_id",
        "父商品ID": "params.parent_product_id",
        "新标题": "params.new_title",
    },
    "发布换图商品": {
        "店铺ID": "shop_id",
        "父商品ID": "params.parent_product_id",
        "新标题": "params.new_title",
        "图片路径": "params.image_path",
    },
}
科学计数法正则 = re.compile(r"^[+-]?\d+\.?\d*[eE][+\-]?\d+$")
整数值正则 = re.compile(r"^\d+$")
整数浮点正则 = re.compile(r"^\d+\.0+$")


class 任务参数服务:
    """task_params 业务服务。"""

    def _校验状态(self, 状态: str) -> None:
        if 状态 not in 允许状态集合:
            raise ValueError(f"不支持的状态: {状态}")

    def _解析状态筛选(self, 状态: Optional[str]) -> List[str]:
        if not 状态:
            return []

        状态列表 = [项.strip() for 项 in str(状态).split(",") if 项.strip()]
        for 当前状态 in 状态列表:
            self._校验状态(当前状态)
        return 状态列表

    def _序列化JSON(self, 数据: Optional[Dict[str, Any]]) -> str:
        return json.dumps(数据 or {}, ensure_ascii=False)

    def _解析JSON(self, 数据: Any) -> Dict[str, Any]:
        if not 数据:
            return {}
        if isinstance(数据, dict):
            return 数据
        try:
            return json.loads(数据)
        except (TypeError, json.JSONDecodeError):
            return {}

    def _转换记录(self, 行: Dict[str, Any]) -> Dict[str, Any]:
        记录 = dict(行)
        记录["params"] = self._解析JSON(记录.get("params"))
        记录["result"] = self._解析JSON(记录.get("result"))
        记录["enabled"] = bool(记录.get("enabled", 1))
        记录["run_count"] = int(记录.get("run_count") or 0)
        return 记录

    async def _店铺是否存在(self, 店铺ID: str) -> bool:
        async with 获取连接() as 连接:
            async with 连接.execute(
                "SELECT 1 FROM shops WHERE id = ?",
                (店铺ID,),
            ) as 游标:
                return await 游标.fetchone() is not None

    async def _根据店铺名称获取ID(self, 店铺名称: str) -> Optional[str]:
        async with 获取连接() as 连接:
            async with 连接.execute(
                "SELECT id FROM shops WHERE name = ? ORDER BY created_at DESC LIMIT 1",
                (店铺名称,),
            ) as 游标:
                结果 = await 游标.fetchone()
                if not 结果:
                    return None
                return str(结果["id"])

    async def _解析店铺标识(self, 店铺标识: str, 行号: int) -> str:
        if 店铺标识.isdigit():
            if not await self._店铺是否存在(店铺标识):
                raise ValueError(f"第 {行号} 行店铺ID不存在: {店铺标识}")
            return 店铺标识

        if await self._店铺是否存在(店铺标识):
            return 店铺标识

        店铺ID = await self._根据店铺名称获取ID(店铺标识)
        if 店铺ID:
            return 店铺ID

        raise ValueError(f"第 {行号} 行店铺名称未找到: {店铺标识}")

    def _解码CSV文本(self, 文件内容: bytes) -> str:
        for 编码 in ("utf-8-sig", "utf-8", "gbk"):
            try:
                return 文件内容.decode(编码)
            except (UnicodeDecodeError, ValueError):
                continue
        raise ValueError("CSV 文件编码不支持，请另存为 UTF-8 格式")

    def _修复科学计数法(self, 值: str) -> str:
        标准值 = str(值).strip()
        if not 标准值:
            return 标准值
        if 科学计数法正则.match(标准值):
            return str(int(float(标准值)))
        return 标准值

    def _预处理CSV行(self, 行数据: Dict[str, str]) -> Dict[str, str]:
        结果: Dict[str, str] = {}
        for 列名, 原值 in 行数据.items():
            字段值 = "" if 原值 is None else str(原值)
            if 列名 and any(标记 in 列名 for 标记 in ("ID", "id", "Id")):
                结果[列名] = self._修复科学计数法(字段值)
            else:
                结果[列名] = 字段值
        return 结果

    def _解析CSV内容(self, 文件内容: bytes) -> List[Dict[str, str]]:
        文本内容 = self._解码CSV文本(文件内容)

        读取器 = csv.DictReader(io.StringIO(文本内容))
        if not 读取器.fieldnames:
            raise ValueError("CSV 文件缺少表头")
        return [self._预处理CSV行(dict(行)) for 行 in 读取器]

    def _解析XLSX内容(self, 文件内容: bytes) -> List[Dict[str, str]]:
        """读取 xlsx 内容并转换为与 CSV 一致的行结构。"""
        工作簿 = load_workbook(io.BytesIO(文件内容), read_only=True)
        try:
            工作表 = 工作簿.worksheets[0]
            行迭代器 = 工作表.iter_rows()
            表头行 = next(行迭代器, None)
            if not 表头行:
                raise ValueError("XLSX 文件缺少表头")

            表头 = ["" if 单元格.value is None else str(单元格.value).strip() for 单元格 in 表头行]
            if not any(表头):
                raise ValueError("XLSX 文件缺少表头")

            结果列表: List[Dict[str, str]] = []
            for 数据行 in 行迭代器:
                行数据: Dict[str, str] = {}
                for 索引, 单元格 in enumerate(数据行):
                    列名 = 表头[索引] if 索引 < len(表头) else ""
                    if not 列名:
                        continue

                    if 单元格.value is None:
                        行数据[列名] = ""
                        continue

                    if (
                        单元格.data_type == "n"
                        and 单元格.value > 9999999999
                    ):
                        行数据[列名] = str(int(单元格.value))
                    else:
                        行数据[列名] = str(单元格.value)

                if any(str(值).strip() for 值 in 行数据.values()):
                    结果列表.append(self._预处理CSV行(行数据))

            return 结果列表
        finally:
            工作簿.close()

    def _解析发布次数(self, 行数据: Dict[str, str], 行号: int) -> int:
        原始值 = str(行数据.get("发布次数", "")).strip()
        if not 原始值:
            return 1

        if 整数值正则.match(原始值):
            发布次数 = int(原始值)
        elif 整数浮点正则.match(原始值):
            发布次数 = int(float(原始值))
        else:
            raise ValueError(f"第 {行号} 行发布次数必须是正整数")

        if 发布次数 <= 0:
            raise ValueError(f"第 {行号} 行发布次数必须大于 0")
        return 发布次数

    async def _映射CSV行(
        self,
        任务名称: str,
        行数据: Dict[str, str],
        行号: int,
    ) -> List[Dict[str, Any]]:
        映射 = CSV映射配置.get(任务名称)
        if not 映射:
            raise ValueError(f"暂不支持的任务类型: {任务名称}")

        缺失列 = [列名 for 列名 in 映射 if 列名 not in 行数据]
        if 缺失列:
            raise ValueError(f"CSV 缺少列: {', '.join(缺失列)}")

        店铺标识 = str(行数据.get("店铺ID", "")).strip()
        父商品ID = str(行数据.get("父商品ID", "")).strip()
        if not 店铺标识:
            raise ValueError(f"第 {行号} 行店铺ID不能为空")
        if not 父商品ID:
            raise ValueError(f"第 {行号} 行父商品ID不能为空")

        店铺ID = await self._解析店铺标识(店铺标识, 行号)

        参数: Dict[str, Any] = {"parent_product_id": 父商品ID}

        新标题 = str(行数据.get("新标题", "")).strip()
        if 新标题:
            参数["new_title"] = 新标题

        图片路径 = str(行数据.get("图片路径", "")).strip()
        if 图片路径:
            参数["image_path"] = 图片路径

        发布次数 = self._解析发布次数(行数据, 行号)
        记录列表: List[Dict[str, Any]] = []
        for 批次序号 in range(1, 发布次数 + 1):
            当前参数 = dict(参数)
            if 发布次数 > 1:
                当前参数["batch_index"] = 批次序号

            记录列表.append(
                {
                    "shop_id": 店铺ID,
                    "task_name": 任务名称,
                    "params": 当前参数,
                    "status": "pending",
                    "result": {},
                    "error": None,
                    "batch_id": None,
                    "enabled": True,
                }
            )

        return 记录列表

    def _构建条件SQL(
        self,
        shop_id: Optional[str] = None,
        task_name: Optional[str] = None,
        status: Optional[str] = None,
        状态集合: Optional[List[str]] = None,
        batch_id: Optional[str] = None,
    ) -> Tuple[str, List[Any], bool]:
        条件列表: List[str] = []
        参数列表: List[Any] = []
        是否包含筛选 = False

        if shop_id:
            条件列表.append("shop_id = ?")
            参数列表.append(shop_id)
            是否包含筛选 = True
        if task_name:
            条件列表.append("task_name = ?")
            参数列表.append(task_name)
            是否包含筛选 = True
        if batch_id:
            条件列表.append("batch_id = ?")
            参数列表.append(batch_id)
            是否包含筛选 = True

        if 状态集合:
            占位符 = ", ".join("?" for _ in 状态集合)
            条件列表.append(f"status IN ({占位符})")
            参数列表.extend(状态集合)
            是否包含筛选 = True
        elif status:
            状态列表 = self._解析状态筛选(status)
            if len(状态列表) == 1:
                条件列表.append("status = ?")
                参数列表.append(状态列表[0])
                是否包含筛选 = True
            elif 状态列表:
                占位符 = ", ".join("?" for _ in 状态列表)
                条件列表.append(f"status IN ({占位符})")
                参数列表.extend(状态列表)
                是否包含筛选 = True

        条件SQL = f"WHERE {' AND '.join(条件列表)}" if 条件列表 else ""
        return 条件SQL, 参数列表, 是否包含筛选

    async def 分页查询(
        self,
        page: int = 1,
        page_size: int = 20,
        shop_id: Optional[str] = None,
        task_name: Optional[str] = None,
        status: Optional[str] = None,
        batch_id: Optional[str] = None,
        updated_from: Optional[str] = None,
        updated_to: Optional[str] = None,
        sort_by: str = "created_at",
        sort_order: str = "desc",
    ) -> Dict[str, Any]:
        """分页查询任务参数记录。"""
        page = max(page, 1)
        page_size = max(page_size, 1)

        条件列表: List[str] = []
        参数列表: List[Any] = []
        状态列表 = self._解析状态筛选(status)

        if shop_id:
            条件列表.append("tp.shop_id = ?")
            参数列表.append(shop_id)
        if task_name:
            条件列表.append("tp.task_name = ?")
            参数列表.append(task_name)
        if batch_id:
            条件列表.append("tp.batch_id = ?")
            参数列表.append(batch_id)
        if len(状态列表) == 1:
            条件列表.append("tp.status = ?")
            参数列表.append(状态列表[0])
        elif 状态列表:
            占位符 = ", ".join("?" for _ in 状态列表)
            条件列表.append(f"tp.status IN ({占位符})")
            参数列表.extend(状态列表)
        if updated_from:
            条件列表.append("DATE(tp.updated_at) >= DATE(?)")
            参数列表.append(updated_from)
        if updated_to:
            条件列表.append("DATE(tp.updated_at) <= DATE(?)")
            参数列表.append(updated_to)

        条件SQL = f"WHERE {' AND '.join(条件列表)}" if 条件列表 else ""
        排序字段映射 = {
            "created_at": "tp.created_at",
            "updated_at": "tp.updated_at",
            "id": "tp.id",
        }
        排序字段 = 排序字段映射.get(sort_by, "tp.created_at")
        排序方向 = "ASC" if str(sort_order).lower() == "asc" else "DESC"

        async with 获取连接() as 连接:
            async with 连接.execute(
                f"""
                SELECT COUNT(*)
                FROM task_params tp
                {条件SQL}
                """,
                参数列表,
            ) as 游标:
                总数行 = await 游标.fetchone()
                总数 = 总数行[0] if 总数行 else 0

            偏移量 = (page - 1) * page_size
            查询参数 = [*参数列表, page_size, 偏移量]
            async with 连接.execute(
                f"""
                SELECT tp.*, s.name AS shop_name
                FROM task_params tp
                LEFT JOIN shops s ON s.id = tp.shop_id
                {条件SQL}
                ORDER BY {排序字段} {排序方向}, tp.id DESC
                LIMIT ? OFFSET ?
                """,
                查询参数,
            ) as 游标:
                行列表 = await 游标.fetchall()

        记录列表 = [self._转换记录(行) for 行 in 行列表]
        return {
            "list": 记录列表,
            "total": 总数,
            "page": page,
            "page_size": page_size,
        }

    async def 获取批次选项(
        self,
        shop_id: Optional[str] = None,
        task_name: Optional[str] = None,
        status: Optional[str] = None,
    ) -> List[Dict[str, Any]]:
        """按当前筛选条件聚合可选批次。"""
        条件列表 = ["tp.batch_id IS NOT NULL", "tp.batch_id != ''"]
        参数列表: List[Any] = []
        状态列表 = self._解析状态筛选(status)

        if shop_id:
            条件列表.append("tp.shop_id = ?")
            参数列表.append(shop_id)
        if task_name:
            条件列表.append("tp.task_name = ?")
            参数列表.append(task_name)
        if len(状态列表) == 1:
            条件列表.append("tp.status = ?")
            参数列表.append(状态列表[0])
        elif 状态列表:
            占位符 = ", ".join("?" for _ in 状态列表)
            条件列表.append(f"tp.status IN ({占位符})")
            参数列表.extend(状态列表)

        条件SQL = f"WHERE {' AND '.join(条件列表)}"

        async with 获取连接() as 连接:
            async with 连接.execute(
                f"""
                SELECT
                    tp.batch_id,
                    COUNT(*) AS record_count,
                    MAX(COALESCE(tp.updated_at, tp.created_at)) AS latest_updated_at
                FROM task_params tp
                {条件SQL}
                GROUP BY tp.batch_id
                ORDER BY record_count DESC, latest_updated_at DESC, tp.batch_id DESC
                """,
                参数列表,
            ) as 游标:
                行列表 = await 游标.fetchall()

        return [
            {
                "batch_id": 行["batch_id"],
                "record_count": int(行["record_count"] or 0),
                "latest_updated_at": 行["latest_updated_at"],
            }
            for 行 in 行列表
        ]

    async def 创建(self, 数据: Dict[str, Any]) -> Dict[str, Any]:
        """创建单条任务参数记录。"""
        店铺ID = str(数据.get("shop_id", "")).strip()
        任务名称 = str(数据.get("task_name", "")).strip()
        状态 = str(数据.get("status") or "pending").strip()
        是否启用 = bool(数据.get("enabled", True))
        执行次数 = int(数据.get("run_count") or 0)
        if not 店铺ID:
            raise ValueError("shop_id 不能为空")
        if not 任务名称:
            raise ValueError("task_name 不能为空")
        self._校验状态(状态)
        if 执行次数 < 0:
            raise ValueError("run_count 不能小于 0")

        if not await self._店铺是否存在(店铺ID):
            raise ValueError("店铺不存在")

        async with 获取连接() as 连接:
            游标 = await 连接.execute(
                """
                INSERT INTO task_params (
                    shop_id, task_name, params, status, result, error, batch_id, enabled, run_count
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    店铺ID,
                    任务名称,
                    self._序列化JSON(数据.get("params")),
                    状态,
                    self._序列化JSON(数据.get("result")),
                    数据.get("error"),
                    数据.get("batch_id"),
                    1 if 是否启用 else 0,
                    执行次数,
                ),
            )
            await 连接.commit()
            记录ID = 游标.lastrowid

        return await self.根据ID获取(记录ID)

    async def 根据ID获取(self, 记录ID: int) -> Optional[Dict[str, Any]]:
        """根据主键获取单条记录。"""
        async with 获取连接() as 连接:
            async with 连接.execute(
                """
                SELECT tp.*, s.name AS shop_name
                FROM task_params tp
                LEFT JOIN shops s ON s.id = tp.shop_id
                WHERE tp.id = ?
                """,
                (记录ID,),
            ) as 游标:
                行 = await 游标.fetchone()

        return self._转换记录(行) if 行 else None

    async def 更新(self, 记录ID: int, 数据: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        """更新单条记录。"""
        现有记录 = await self.根据ID获取(记录ID)
        if not 现有记录:
            return None

        更新字段: List[str] = []
        更新值: List[Any] = []

        if "shop_id" in 数据 and 数据["shop_id"] is not None:
            店铺ID = str(数据["shop_id"]).strip()
            if not 店铺ID:
                raise ValueError("shop_id 不能为空")
            if not await self._店铺是否存在(店铺ID):
                raise ValueError("店铺不存在")
            更新字段.append("shop_id = ?")
            更新值.append(店铺ID)

        if "task_name" in 数据 and 数据["task_name"] is not None:
            任务名称 = str(数据["task_name"]).strip()
            if not 任务名称:
                raise ValueError("task_name 不能为空")
            更新字段.append("task_name = ?")
            更新值.append(任务名称)

        if "params" in 数据 and 数据["params"] is not None:
            更新字段.append("params = ?")
            更新值.append(self._序列化JSON(数据["params"]))

        if "status" in 数据 and 数据["status"] is not None:
            状态 = str(数据["status"]).strip()
            self._校验状态(状态)
            更新字段.append("status = ?")
            更新值.append(状态)

        if "result" in 数据 and 数据["result"] is not None:
            更新字段.append("result = ?")
            更新值.append(self._序列化JSON(数据["result"]))

        if "error" in 数据:
            更新字段.append("error = ?")
            更新值.append(数据.get("error"))

        if "batch_id" in 数据:
            更新字段.append("batch_id = ?")
            更新值.append(数据.get("batch_id"))

        if "enabled" in 数据 and 数据["enabled"] is not None:
            更新字段.append("enabled = ?")
            更新值.append(1 if bool(数据["enabled"]) else 0)

        if not 更新字段:
            return 现有记录

        更新字段.append("updated_at = CURRENT_TIMESTAMP")
        更新值.append(记录ID)

        async with 获取连接() as 连接:
            await 连接.execute(
                f"UPDATE task_params SET {', '.join(更新字段)} WHERE id = ?",
                更新值,
            )
            await 连接.commit()

        return await self.根据ID获取(记录ID)

    async def 删除(self, 记录ID: int) -> bool:
        """删除单条记录。"""
        async with 获取连接() as 连接:
            游标 = await 连接.execute("DELETE FROM task_params WHERE id = ?", (记录ID,))
            await 连接.commit()
            return 游标.rowcount > 0

    async def 按条件清空(
        self,
        shop_id: Optional[str] = None,
        task_name: Optional[str] = None,
        status: Optional[str] = None,
        batch_id: Optional[str] = None,
    ) -> int:
        """按条件删除多条记录并返回删除数量。"""
        条件SQL, 参数列表, _ = self._构建条件SQL(
            shop_id=shop_id,
            task_name=task_name,
            status=status,
            batch_id=batch_id,
        )

        async with 获取连接() as 连接:
            游标 = await 连接.execute(
                f"DELETE FROM task_params {条件SQL}",
                参数列表,
            )
            await 连接.commit()
            return 游标.rowcount

    async def 批量导入(
        self,
        文件内容: bytes,
        task_name: str,
        file_name: str = "",
    ) -> Dict[str, Any]:
        """解析 CSV 或 XLSX 并批量导入。"""
        文件后缀 = Path(file_name or "").suffix.lower()
        if 文件后缀 == ".xlsx":
            行列表 = self._解析XLSX内容(文件内容)
        else:
            行列表 = self._解析CSV内容(文件内容)
        成功数量 = 0
        失败数量 = 0
        错误列表: List[str] = []

        for 行号, 行数据 in enumerate(行列表, start=2):
            try:
                记录列表 = await self._映射CSV行(task_name, 行数据, 行号)
                for 记录 in 记录列表:
                    await self.创建(记录)
                    成功数量 += 1
            except Exception as 异常:
                失败数量 += 1
                错误列表.append(str(异常))

        return {
            "success_count": 成功数量,
            "failed_count": 失败数量,
            "errors": 错误列表,
        }

    async def 获取待执行列表(self, shop_id: str, task_name: str) -> List[Dict[str, Any]]:
        """获取指定店铺和任务的待执行记录。"""
        async with 获取连接() as 连接:
            async with 连接.execute(
                """
                SELECT tp.*, s.name AS shop_name
                FROM task_params tp
                LEFT JOIN shops s ON s.id = tp.shop_id
                WHERE tp.shop_id = ? AND tp.task_name = ? AND tp.status = 'pending' AND tp.enabled = 1
                ORDER BY tp.id ASC
                """,
                (shop_id, task_name),
            ) as 游标:
                行列表 = await 游标.fetchall()

        return [self._转换记录(行) for 行 in 行列表]

    async def 查询批次成功记录(
        self,
        shop_id: str,
        batch_id: str,
        task_name: str,
    ) -> List[Dict[str, Any]]:
        """查询同批次中指定任务的成功结果 JSON 列表。"""
        async with 获取连接() as 连接:
            async with 连接.execute(
                """
                SELECT result
                FROM task_params
                WHERE shop_id = ?
                  AND batch_id = ?
                  AND task_name = ?
                  AND status = 'success'
                ORDER BY id ASC
                """,
                (shop_id, batch_id, task_name),
            ) as 游标:
                行列表 = await 游标.fetchall()

        return [self._解析JSON(行["result"]) for 行 in 行列表]

    async def 启用(self, 记录ID: int) -> Optional[Dict[str, Any]]:
        """启用单条任务参数记录。"""
        return await self.更新(记录ID, {"enabled": True})

    async def 禁用(self, 记录ID: int) -> Optional[Dict[str, Any]]:
        """禁用单条任务参数记录。"""
        return await self.更新(记录ID, {"enabled": False})

    async def 重置(self, 记录ID: int) -> Optional[Dict[str, Any]]:
        """重置单条任务参数记录到 pending。"""
        现有记录 = await self.根据ID获取(记录ID)
        if not 现有记录:
            return None

        async with 获取连接() as 连接:
            await 连接.execute(
                """
                UPDATE task_params
                SET status = 'pending',
                    result = '{}',
                    error = NULL,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (记录ID,),
            )
            await 连接.commit()

        return await self.根据ID获取(记录ID)

    async def _批量更新(
        self,
        更新SQL: str,
        参数列表: List[Any],
    ) -> int:
        async with 获取连接() as 连接:
            游标 = await 连接.execute(更新SQL, 参数列表)
            await 连接.commit()
            return 游标.rowcount

    async def 批量重置(
        self,
        shop_id: Optional[str] = None,
        task_name: Optional[str] = None,
        status: Optional[str] = None,
        batch_id: Optional[str] = None,
    ) -> int:
        """按筛选条件批量重置记录。"""
        if status:
            目标状态集合 = self._解析状态筛选(status)
        else:
            目标状态集合 = ["success", "failed"]

        条件SQL, 参数列表, _ = self._构建条件SQL(
            shop_id=shop_id,
            task_name=task_name,
            状态集合=目标状态集合,
            batch_id=batch_id,
        )
        if not 条件SQL:
            raise ValueError("批量重置必须至少包含一个筛选条件")

        return await self._批量更新(
            f"""
            UPDATE task_params
            SET status = 'pending',
                result = '{{}}',
                error = NULL,
                updated_at = CURRENT_TIMESTAMP
            {条件SQL}
            """,
            参数列表,
        )

    async def 批量启用(
        self,
        shop_id: Optional[str] = None,
        task_name: Optional[str] = None,
        status: Optional[str] = None,
        batch_id: Optional[str] = None,
    ) -> int:
        """按筛选条件批量启用记录。"""
        条件SQL, 参数列表, 是否包含筛选 = self._构建条件SQL(
            shop_id=shop_id,
            task_name=task_name,
            status=status,
            batch_id=batch_id,
        )
        if not 是否包含筛选:
            raise ValueError("批量启用必须至少提供一个筛选条件")

        return await self._批量更新(
            f"""
            UPDATE task_params
            SET enabled = 1,
                updated_at = CURRENT_TIMESTAMP
            {条件SQL}
            """,
            参数列表,
        )

    async def 批量禁用(
        self,
        shop_id: Optional[str] = None,
        task_name: Optional[str] = None,
        status: Optional[str] = None,
        batch_id: Optional[str] = None,
    ) -> int:
        """按筛选条件批量禁用记录。"""
        条件SQL, 参数列表, 是否包含筛选 = self._构建条件SQL(
            shop_id=shop_id,
            task_name=task_name,
            status=status,
            batch_id=batch_id,
        )
        if not 是否包含筛选:
            raise ValueError("批量禁用必须至少提供一个筛选条件")

        return await self._批量更新(
            f"""
            UPDATE task_params
            SET enabled = 0,
                updated_at = CURRENT_TIMESTAMP
            {条件SQL}
            """,
            参数列表,
        )

    async def 更新执行结果(
        self,
        记录ID: int,
        状态: str,
        结果: Optional[Dict[str, Any]] = None,
        错误信息: Optional[str] = None,
    ) -> Optional[Dict[str, Any]]:
        """回填执行结果。"""
        self._校验状态(状态)

        更新字段 = [
            "status = ?",
            "result = ?",
            "error = ?",
            "updated_at = CURRENT_TIMESTAMP",
        ]
        更新参数: List[Any] = [
            状态,
            self._序列化JSON(结果),
            错误信息,
        ]

        if 状态 in {"success", "failed"}:
            更新字段.append("run_count = COALESCE(run_count, 0) + 1")

        async with 获取连接() as 连接:
            await 连接.execute(
                f"""
                UPDATE task_params
                SET {', '.join(更新字段)}
                WHERE id = ?
                """,
                [*更新参数, 记录ID],
            )
            await 连接.commit()

        return await self.根据ID获取(记录ID)


任务参数服务实例 = 任务参数服务()
